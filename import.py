"""
Unified import script for Monta Glossary.

This script handles:
1. Importing data from Excel to SQLite database
2. Applying amendments and alternatives
3. Generating markdown file from database
"""

import os
import sys
import time
import json
import argparse
from openpyxl import load_workbook
from dotenv import load_dotenv

# Add python package directory to path
project_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(project_dir, 'python'))

from monta_glossary.database import GlossaryDB
from monta_glossary import Glossary


# ============================================================================
# PART 1: IMPORT FROM EXCEL
# ============================================================================

def parse_alternative_words(row, columns):
    """Parse alternative words from the row if they exist."""
    # Look for 'Alternative words' column
    alt_col_index = None
    for i, col in enumerate(columns):
        if col and 'alternative' in col.lower():
            alt_col_index = i
            break

    if alt_col_index is None or not row[alt_col_index]:
        return []

    # Parse the alternative words (usually comma-separated)
    alt_text = str(row[alt_col_index])
    alternatives = [alt.strip() for alt in alt_text.split(',')]
    return [alt for alt in alternatives if alt]


def import_from_excel(excel_path: str, db_path: str):
    """Import glossary data from Excel file to SQLite database."""
    if not os.path.exists(excel_path):
        print(f"❌ Error: Excel file not found at {excel_path}")
        sys.exit(1)

    print(f"📂 Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Get column headers
    columns = [cell.value for cell in ws[1]]

    # Create column index mapping
    col_idx = {col: i for i, col in enumerate(columns) if col}

    # Language columns
    language_columns = ['en', 'en_US', 'da', 'de', 'fr', 'no', 'ro', 'sv', 'es', 'it', 'nl', 'fr_CA']

    # Initialize database
    print(f"🗄️  Initializing database: {db_path}")
    db = GlossaryDB(db_path)
    db.create_tables()

    # Clear existing data
    print("🧹 Clearing existing data...")
    db.clear_all_data()

    # Import data
    print("📥 Importing terms...")
    imported_count = 0
    skipped_count = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Get term name
        term = row[col_idx['term']] if 'term' in col_idx else None

        if not term:
            skipped_count += 1
            continue

        # Get metadata
        description = row[col_idx['description']] if 'description' in col_idx else ''
        case_sensitive = str(row[col_idx['casesensitive']]).lower() == 'yes' if 'casesensitive' in col_idx else False
        translatable = str(row[col_idx['translatable']]).lower() == 'yes' if 'translatable' in col_idx else True
        forbidden = str(row[col_idx['forbidden']]).lower() == 'yes' if 'forbidden' in col_idx else False
        tags = row[col_idx['tags']] if 'tags' in col_idx else None

        # Insert term
        try:
            term_id = db.insert_term(
                term=term,
                description=description if description else '',
                case_sensitive=case_sensitive,
                translatable=translatable,
                forbidden=forbidden,
                tags=tags
            )

            # Insert translations
            for lang in language_columns:
                if lang in col_idx and row[col_idx[lang]]:
                    translation = row[col_idx[lang]]
                    db.insert_translation(term_id, lang, translation)

            # Insert alternative words if present
            alternatives = parse_alternative_words(row, columns)
            for alt in alternatives:
                db.insert_alternative_word(term_id, alt)

            # Insert additional descriptions
            if 'nl_description' in col_idx and row[col_idx['nl_description']]:
                db.insert_additional_description(term_id, 'nl', row[col_idx['nl_description']])

            if 'en_US_description' in col_idx and row[col_idx['en_US_description']]:
                db.insert_additional_description(term_id, 'en_US', row[col_idx['en_US_description']])

            imported_count += 1

            if imported_count % 100 == 0:
                print(f"  ✓ Imported {imported_count} terms...")

        except Exception as e:
            print(f"  ⚠️  Error importing term '{term}' at row {row_num}: {e}")
            skipped_count += 1

    db.close()

    print(f"\n✅ Import complete!")
    print(f"   • Successfully imported: {imported_count} terms")
    if skipped_count > 0:
        print(f"   • Skipped: {skipped_count} rows")

    return imported_count


# ============================================================================
# PART 2: GENERATE ALTERNATIVE WORDS (OPENAI)
# ============================================================================

def apply_alternatives_from_json(db_path: str, alternatives_path: str) -> dict:
    """Apply alternatives from alternatives.json to database.

    Args:
        db_path: Path to SQLite database
        alternatives_path: Path to alternatives.json file

    Returns:
        Dictionary with statistics
    """
    import inflect
    p = inflect.engine()

    if not os.path.exists(alternatives_path):
        print(f"⚠️  No alternatives file found at {alternatives_path}")
        return {'total': 0, 'applied': 0, 'skipped': 0}

    print(f"📝 Applying alternatives from: {alternatives_path}")

    # Load alternatives.json
    try:
        with open(alternatives_path, 'r', encoding='utf-8') as f:
            alternatives_data = json.load(f)
    except Exception as e:
        print(f"❌ Error loading alternatives file: {e}")
        return {'total': 0, 'applied': 0, 'skipped': 0}

    alternatives_map = alternatives_data.get('alternatives', {})

    if not alternatives_map:
        print("⚠️  No alternatives to apply")
        return {'total': 0, 'applied': 0, 'skipped': 0}

    # Initialize database
    db = GlossaryDB(db_path)

    stats = {
        'total': len(alternatives_map),
        'applied': 0,
        'skipped': 0
    }

    print(f"Processing {len(alternatives_map)} term(s) with alternatives...")

    for term_name, alternatives_list in alternatives_map.items():
        # Get term (case-insensitive)
        term_data = get_term_by_name_case_insensitive(db, term_name)

        if not term_data:
            print(f"  ⚠️  Term '{term_name}' not found in database")
            stats['skipped'] += 1
            continue

        term_id = term_data['id']

        # Get existing alternatives from database
        existing_alternatives = db.get_alternative_words(term_id)

        # Add new alternatives (avoid duplicates)
        added = 0
        for alt in alternatives_list:
            if alt not in existing_alternatives:
                # Auto-detect if this alternative is a plural form
                # Check the last word in multi-word terms
                words = alt.split()
                last_word = words[-1] if words else alt

                # Use inflect to check if it's plural
                is_plural_alt = p.singular_noun(last_word) is not False

                db.insert_alternative_word(term_id, alt, is_plural=is_plural_alt)
                added += 1

        if added > 0:
            stats['applied'] += 1

    db.close()

    print(f"\n✅ Alternatives applied!")
    print(f"   • Total terms: {stats['total']}")
    print(f"   • Applied: {stats['applied']}")
    print(f"   • Skipped: {stats['skipped']}")

    return stats


def apply_plurals_from_json(db_path: str, plurals_path: str) -> dict:
    """Apply plurals from plurals.json to database as alternative words.

    Args:
        db_path: Path to SQLite database
        plurals_path: Path to plurals.json file

    Returns:
        Dictionary with statistics
    """
    if not os.path.exists(plurals_path):
        print(f"⚠️  No plurals file found at {plurals_path}")
        return {'total': 0, 'applied': 0, 'skipped': 0}

    print(f"📝 Applying plurals from: {plurals_path}")

    # Load plurals.json
    try:
        with open(plurals_path, 'r', encoding='utf-8') as f:
            plurals_data = json.load(f)
    except Exception as e:
        print(f"❌ Error loading plurals file: {e}")
        return {'total': 0, 'applied': 0, 'skipped': 0}

    plurals_map = plurals_data.get('plurals', {})

    if not plurals_map:
        print("⚠️  No plurals to apply")
        return {'total': 0, 'applied': 0, 'skipped': 0}

    # Initialize database
    db = GlossaryDB(db_path)

    stats = {
        'total': len(plurals_map),
        'applied': 0,
        'skipped': 0
    }

    print(f"Processing {len(plurals_map)} term(s) with plurals...")

    for term_name, plural_info in plurals_map.items():
        # Get term (case-insensitive)
        term_data = get_term_by_name_case_insensitive(db, term_name)

        if not term_data:
            # Term might not be in database yet (e.g., from external_glossary.md)
            stats['skipped'] += 1
            continue

        term_id = term_data['id']

        # Get the plural form
        plural = plural_info.get('plural')
        singular = plural_info.get('singular')

        if not plural or not singular:
            print(f"  ⚠️  Invalid plural entry for '{term_name}'")
            stats['skipped'] += 1
            continue

        # Update the term's plural_form column
        db.update_term_plural(term_id, plural)

        # Get existing alternatives from database
        existing_alternatives = db.get_alternative_words(term_id)

        # Add both singular and plural as alternatives (if not already present)
        added = 0

        # Add plural form as alternative (marked as is_plural=True)
        if plural != term_name and plural not in existing_alternatives:
            db.insert_alternative_word(term_id, plural, is_plural=True)
            added += 1

        # Add singular form as alternative (useful if the canonical term is different)
        if singular != term_name and singular not in existing_alternatives:
            db.insert_alternative_word(term_id, singular, is_plural=False)
            added += 1

        if added > 0:
            stats['applied'] += 1
        else:
            stats['skipped'] += 1

    db.close()

    print(f"\n✅ Plurals applied!")
    print(f"   • Total terms: {stats['total']}")
    print(f"   • Applied: {stats['applied']}")
    print(f"   • Skipped: {stats['skipped']}")

    return stats


# ============================================================================
# PART 3: GENERATE MARKDOWN
# ============================================================================

def format_term_as_markdown(term) -> str:
    """Format a Term object as markdown."""
    lines = []

    # Term heading
    lines.append(f"## {term.term}")

    # Tags
    if term.tags:
        lines.append(f"**Tag:** {term.tags}")
        lines.append("")

    # Alternative words
    if term.alternative_words:
        alt_words = ", ".join(term.alternative_words)
        lines.append(f"**Alternative words:** {alt_words}")
        lines.append("")

    # Description
    lines.append(f"**Description:** {term.description}")
    lines.append("")

    # Translations
    if term.translations:
        lines.append("**Translations:**")
        # Sort languages for consistent output
        for lang_code in sorted(term.translations.keys()):
            translation = term.translations[lang_code]
            lines.append(f"- {lang_code.upper()}: {translation}")
        lines.append("")

    # Separator
    lines.append("---")
    lines.append("")

    return "\n".join(lines)


def generate_markdown(output_path: str, db_path: str):
    """Generate glossary.md file from database."""
    print(f"📝 Generating markdown file: {output_path}")

    # Initialize glossary
    with Glossary(db_path) as glossary:
        # Get all terms
        all_terms = glossary.get_all()

        if not all_terms:
            print("⚠️  Warning: No terms found in database!")
            return 0

        # Generate markdown content
        lines = []

        # Header
        lines.append("# Monta Glossary (AI-Ready)")
        lines.append("")
        lines.append("This glossary maps Monta terminology for AI models. Each entry includes description, usage rules, and translations.")
        lines.append("")

        # Add each term
        for i, term in enumerate(all_terms, 1):
            lines.append(format_term_as_markdown(term))

            if i % 100 == 0:
                print(f"  ✓ Processed {i}/{len(all_terms)} terms...")

        # Write to file
        content = "\n".join(lines)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)

        print(f"\n✅ Markdown generation complete!")
        print(f"   • Total terms: {len(all_terms)}")
        print(f"   • File size: {len(content):,} characters")

        return len(all_terms)


# ============================================================================
# PART 4: APPLY AMENDMENTS
# ============================================================================

def get_term_by_name_case_insensitive(db: GlossaryDB, term_name: str) -> dict:
    """Get a term by name, case-insensitive.

    Args:
        db: Database instance
        term_name: Term name to search for (case-insensitive)

    Returns:
        Term data dictionary or None if not found
    """
    cursor = db.conn.cursor()
    cursor.execute("""
        SELECT * FROM terms
        WHERE LOWER(term) = LOWER(?)
        LIMIT 1
    """, (term_name,))
    row = cursor.fetchone()
    return dict(row) if row else None


def load_amendments(amendments_path: str) -> dict:
    """Load and validate amendments JSON file."""
    if not os.path.exists(amendments_path):
        print(f"⚠️  Warning: Amendments file not found at {amendments_path}")
        return None

    try:
        with open(amendments_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Basic validation
        if 'amendments' not in data:
            print("❌ Error: Invalid amendments file - missing 'amendments' key")
            return None

        if not isinstance(data['amendments'], list):
            print("❌ Error: Invalid amendments file - 'amendments' must be a list")
            return None

        return data

    except json.JSONDecodeError as e:
        print(f"❌ Error: Invalid JSON in amendments file: {e}")
        return None
    except Exception as e:
        print(f"❌ Error loading amendments file: {e}")
        return None


def apply_update_term(db: GlossaryDB, amendment: dict, stats: dict):
    """Apply update_term amendment."""
    term_name = amendment.get('term')
    changes = amendment.get('changes', {})

    if not term_name or not changes:
        print(f"  ⚠️  Skipping invalid update_term amendment")
        stats['skipped'] += 1
        return

    # Get existing term (case-insensitive)
    term_data = get_term_by_name_case_insensitive(db, term_name)
    if not term_data:
        print(f"  ⚠️  Term '{term_name}' not found, cannot update")
        stats['skipped'] += 1
        return

    # Update term fields
    term_id = term_data['id']
    cursor = db.conn.cursor()

    update_fields = []
    update_values = []

    if 'description' in changes:
        update_fields.append('description = ?')
        update_values.append(changes['description'])

    if 'case_sensitive' in changes:
        update_fields.append('case_sensitive = ?')
        update_values.append(1 if changes['case_sensitive'] else 0)

    if 'translatable' in changes:
        update_fields.append('translatable = ?')
        update_values.append(1 if changes['translatable'] else 0)

    if 'forbidden' in changes:
        update_fields.append('forbidden = ?')
        update_values.append(1 if changes['forbidden'] else 0)

    if 'tags' in changes:
        update_fields.append('tags = ?')
        update_values.append(changes['tags'])

    if update_fields:
        update_values.append(term_id)
        query = f"UPDATE terms SET {', '.join(update_fields)} WHERE id = ?"
        cursor.execute(query, update_values)
        db.conn.commit()

        print(f"  ✓ Updated term '{term_name}'")
        stats['updated'] += 1
    else:
        stats['skipped'] += 1


def apply_add_alternatives(db: GlossaryDB, amendment: dict, stats: dict):
    """Apply add_alternatives amendment."""
    term_name = amendment.get('term')
    alternatives = amendment.get('alternatives', [])

    if not term_name or not alternatives:
        print(f"  ⚠️  Skipping invalid add_alternatives amendment")
        stats['skipped'] += 1
        return

    # Get existing term (case-insensitive)
    term_data = get_term_by_name_case_insensitive(db, term_name)
    if not term_data:
        print(f"  ⚠️  Term '{term_name}' not found, cannot add alternatives")
        stats['skipped'] += 1
        return

    term_id = term_data['id']

    # Get existing alternatives to avoid duplicates
    existing = db.get_alternative_words(term_id)

    added = 0
    for alt in alternatives:
        if alt not in existing:
            db.insert_alternative_word(term_id, alt)
            added += 1

    if added > 0:
        print(f"  ✓ Added {added} alternative(s) to '{term_name}'")
        stats['updated'] += 1
    else:
        stats['skipped'] += 1


def apply_remove_alternative(db: GlossaryDB, amendment: dict, stats: dict):
    """Apply remove_alternative amendment."""
    term_name = amendment.get('term')
    alternative = amendment.get('alternative')

    if not term_name or not alternative:
        print(f"  ⚠️  Skipping invalid remove_alternative amendment")
        stats['skipped'] += 1
        return

    # Get existing term (case-insensitive)
    term_data = get_term_by_name_case_insensitive(db, term_name)
    if not term_data:
        print(f"  ⚠️  Term '{term_name}' not found")
        stats['skipped'] += 1
        return

    term_id = term_data['id']
    cursor = db.conn.cursor()
    cursor.execute("""
        DELETE FROM alternative_words
        WHERE term_id = ? AND alternative = ?
    """, (term_id, alternative))
    db.conn.commit()

    if cursor.rowcount > 0:
        print(f"  ✓ Removed alternative '{alternative}' from '{term_name}'")
        stats['updated'] += 1
    else:
        stats['skipped'] += 1


def apply_add_translation(db: GlossaryDB, amendment: dict, stats: dict):
    """Apply add_translation amendment."""
    term_name = amendment.get('term')
    translations = amendment.get('translations', {})

    if not term_name or not translations:
        print(f"  ⚠️  Skipping invalid add_translation amendment")
        stats['skipped'] += 1
        return

    # Get existing term (case-insensitive)
    term_data = get_term_by_name_case_insensitive(db, term_name)
    if not term_data:
        print(f"  ⚠️  Term '{term_name}' not found, cannot add translations")
        stats['skipped'] += 1
        return

    term_id = term_data['id']

    added = 0
    for lang_code, translation in translations.items():
        db.insert_translation(term_id, lang_code, translation)
        added += 1

    if added > 0:
        print(f"  ✓ Added/updated {added} translation(s) for '{term_name}'")
        stats['updated'] += 1
    else:
        stats['skipped'] += 1


def apply_add_description(db: GlossaryDB, amendment: dict, stats: dict):
    """Apply add_description amendment."""
    term_name = amendment.get('term')
    descriptions = amendment.get('descriptions', {})

    if not term_name or not descriptions:
        print(f"  ⚠️  Skipping invalid add_description amendment")
        stats['skipped'] += 1
        return

    # Get existing term (case-insensitive)
    term_data = get_term_by_name_case_insensitive(db, term_name)
    if not term_data:
        print(f"  ⚠️  Term '{term_name}' not found, cannot add descriptions")
        stats['skipped'] += 1
        return

    term_id = term_data['id']

    added = 0
    for lang_code, description in descriptions.items():
        db.insert_additional_description(term_id, lang_code, description)
        added += 1

    if added > 0:
        print(f"  ✓ Added/updated {added} description(s) for '{term_name}'")
        stats['updated'] += 1
    else:
        stats['skipped'] += 1


def apply_add_term(db: GlossaryDB, amendment: dict, stats: dict):
    """Apply add_term amendment."""
    term_name = amendment.get('term')
    data = amendment.get('data', {})

    if not term_name:
        print(f"  ⚠️  Skipping invalid add_term amendment - no term name")
        stats['skipped'] += 1
        return

    # Check if term already exists (case-insensitive)
    existing = get_term_by_name_case_insensitive(db, term_name)
    if existing:
        print(f"  ⚠️  Term '{term_name}' already exists, skipping add")
        stats['skipped'] += 1
        return

    # Insert new term
    try:
        term_id = db.insert_term(
            term=term_name,
            description=data.get('description', ''),
            case_sensitive=data.get('case_sensitive', False),
            translatable=data.get('translatable', True),
            forbidden=data.get('forbidden', False),
            tags=data.get('tags', None)
        )

        # Add translations
        if 'translations' in data:
            for lang_code, translation in data['translations'].items():
                db.insert_translation(term_id, lang_code, translation)

        # Add alternatives
        if 'alternatives' in data:
            for alt in data['alternatives']:
                db.insert_alternative_word(term_id, alt)

        # Add additional descriptions
        if 'descriptions' in data:
            for lang_code, description in data['descriptions'].items():
                db.insert_additional_description(term_id, lang_code, description)

        print(f"  ✓ Added new term '{term_name}'")
        stats['added'] += 1

    except Exception as e:
        print(f"  ⚠️  Error adding term '{term_name}': {e}")
        stats['skipped'] += 1


def apply_delete_term(db: GlossaryDB, amendment: dict, stats: dict):
    """Apply delete_term amendment."""
    term_name = amendment.get('term')

    if not term_name:
        print(f"  ⚠️  Skipping invalid delete_term amendment")
        stats['skipped'] += 1
        return

    # Get existing term (case-insensitive)
    term_data = get_term_by_name_case_insensitive(db, term_name)
    if not term_data:
        print(f"  ⚠️  Term '{term_name}' not found, cannot delete")
        stats['skipped'] += 1
        return

    term_id = term_data['id']
    cursor = db.conn.cursor()

    # Delete term (cascade will handle related records)
    cursor.execute("DELETE FROM terms WHERE id = ?", (term_id,))
    db.conn.commit()

    print(f"  ✓ Deleted term '{term_name}'")
    stats['deleted'] += 1


def apply_amendments(db_path: str, amendments_path: str) -> dict:
    """Apply amendments from JSON file to database.

    Args:
        db_path: Path to SQLite database
        amendments_path: Path to amendments JSON file

    Returns:
        Dictionary with statistics about applied amendments
    """
    print(f"📝 Applying amendments from: {amendments_path}")

    # Load amendments
    amendments_data = load_amendments(amendments_path)
    if not amendments_data:
        return {'total': 0, 'updated': 0, 'added': 0, 'deleted': 0, 'skipped': 0}

    amendments = amendments_data.get('amendments', [])

    if not amendments:
        print("⚠️  No amendments to apply")
        return {'total': 0, 'updated': 0, 'added': 0, 'deleted': 0, 'skipped': 0}

    # Initialize database
    db = GlossaryDB(db_path)

    # Statistics
    stats = {
        'total': len(amendments),
        'updated': 0,
        'added': 0,
        'deleted': 0,
        'skipped': 0
    }

    print(f"Processing {len(amendments)} amendment(s)...")

    # Apply each amendment
    for i, amendment in enumerate(amendments, 1):
        amendment_type = amendment.get('type')

        if not amendment_type:
            print(f"  ⚠️  Amendment #{i}: Missing 'type' field")
            stats['skipped'] += 1
            continue

        # Dispatch to appropriate handler
        if amendment_type == 'update_term':
            apply_update_term(db, amendment, stats)
        elif amendment_type == 'add_alternatives':
            apply_add_alternatives(db, amendment, stats)
        elif amendment_type == 'remove_alternative':
            apply_remove_alternative(db, amendment, stats)
        elif amendment_type == 'add_translation':
            apply_add_translation(db, amendment, stats)
        elif amendment_type == 'add_description':
            apply_add_description(db, amendment, stats)
        elif amendment_type == 'add_term':
            apply_add_term(db, amendment, stats)
        elif amendment_type == 'delete_term':
            apply_delete_term(db, amendment, stats)
        else:
            print(f"  ⚠️  Unknown amendment type: {amendment_type}")
            stats['skipped'] += 1

    db.close()

    print(f"\n✅ Amendments applied!")
    print(f"   • Total amendments: {stats['total']}")
    print(f"   • Updated: {stats['updated']}")
    print(f"   • Added: {stats['added']}")
    print(f"   • Deleted: {stats['deleted']}")
    print(f"   • Skipped: {stats['skipped']}")

    return stats


# ============================================================================
# MAIN
# ============================================================================

def main():
    """Main entry point for the unified import script."""
    parser = argparse.ArgumentParser(
        description='Monta Glossary - Unified Import Tool',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Standard import (Excel → amendments → alternatives → plurals → markdown)
  python import.py

  # Reapply amendments/alternatives/plurals without Excel reimport
  python import.py --skip-import

  # Full clean rebuild (clears database first)
  python import.py --truncate-db

Note:
  - Amendments, alternatives, and plurals are ALWAYS applied automatically
  - To generate AI alternatives: python generate_alternatives.py
  - To regenerate plurals: python generate_plurals.py
        """
    )

    parser.add_argument(
        '--skip-import',
        action='store_true',
        help='Skip Excel import, only run processing pipeline (amendments/alternatives/plurals/markdown)'
    )

    parser.add_argument(
        '--truncate-db',
        action='store_true',
        help='Truncate/clear database before starting (useful for clean rebuild)'
    )

    args = parser.parse_args()

    # Determine paths
    project_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(project_dir, 'files', 'inputs', 'monta_raw_glossary.xlsx')
    amendments_path = os.path.join(project_dir, 'files', 'inputs', 'amendments.json')
    alternatives_path = os.path.join(project_dir, 'files', 'inputs', 'alternatives.json')
    plurals_path = os.path.join(project_dir, 'files', 'inputs', 'plurals.json')
    db_path = os.path.join(project_dir, 'files', 'outputs', 'glossary.sqlite')
    md_path = os.path.join(project_dir, 'files', 'outputs', 'glossary.md')

    print("=" * 70)
    print("  Monta Glossary - Unified Import Tool")
    print("=" * 70)
    print()

    # Step 0: Truncate database if requested
    if args.truncate_db:
        if os.path.exists(db_path):
            print("🗑️  Truncating database...")
            db = GlossaryDB(db_path)
            db.clear_all_data()
            db.close()
            print("   ✓ Database truncated\n")
        else:
            print("⚠️  Database doesn't exist yet, skipping truncation\n")

    # Step 1: Import from Excel
    if not args.skip_import:
        if not os.path.exists(excel_path):
            print(f"❌ Error: Excel file not found at {excel_path}")
            sys.exit(1)

        import_from_excel(excel_path, db_path)
        print()
    else:
        print("⏭️  Skipping Excel import")
        print()
        if not os.path.exists(db_path):
            print(f"❌ Error: Database not found at {db_path}")
            print("Cannot skip import if database doesn't exist!")
            sys.exit(1)

    # Step 2: Apply amendments (always)
    print("=" * 70)
    apply_amendments(db_path, amendments_path)
    print()

    # Step 3: Apply alternatives from alternatives.json (always)
    print("=" * 70)
    apply_alternatives_from_json(db_path, alternatives_path)
    print()

    # Step 4: Apply plurals from plurals.json (always, if exists)
    print("=" * 70)
    apply_plurals_from_json(db_path, plurals_path)
    print()

    # Step 5: Generate markdown (always)
    print("=" * 70)
    generate_markdown(md_path, db_path)
    print()

    # Summary
    print("=" * 70)
    print("✨ All done!")
    print()
    print("📊 Quick stats:")
    with Glossary(db_path) as glossary:
        print(f"   • Total terms: {glossary.count()}")
        print(f"   • Languages: {len(glossary.get_languages())}")

    print()
    print("💡 Next steps:")
    print("   • Edit files/inputs/alternatives.json to manually add/modify alternatives")
    print("   • Run python generate_plurals.py to regenerate plurals.json")
    print("   • Use the glossary in your Python/Kotlin/TypeScript projects")
    print()


if __name__ == '__main__':
    main()
